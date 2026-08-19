"""
Scraper Sekolah Swasta (SD/SMP/SMA/SMK) - data.kemendikdasmen.go.id

Kenapa pake Playwright padahal ambil data via API?
API-nya (api.data.belajar.id) diproteksi WAF yg cek TLS fingerprint koneksi
HTTPS - request polos dari Python (`requests`) ketolak 403 walau header disamain
persis kayak browser. Playwright buka Chromium ASLI, jadi TLS-nya genuine,
WAF lolos. Tapi kita GAK KLIK UI SAMA SEKALI - cuma manggil fetch() dari dalem
browser context (page.evaluate), jadi tetep ngebut kayak API murni.

API dipake:
  LIST   : https://api.data.belajar.id/data-portal-backend/v2/master-data/satuan-pendidikan/daftar-data-induk/{kodeWilayah}
  DETAIL : https://api.data.belajar.id/data-portal-backend/v1/master-data/satuan-pendidikan/details/{npsn}

CARA PAKAI:
    pip install playwright
    playwright install chromium
    python scrape_api.py

    Muncul menu pilih provinsi, ketik nomor (bisa banyak dipisah koma, atau 0 = semua).

OUTPUT:
    {NamaProvinsi}_Swasta.csv  -> 1 file per provinsi
    Kolom: NPSN | Nama Sekolah | Telpon | Email | Akreditasi | Total Peserta Didik |
           Total Pendidik | Naungan | Kabupaten | Kecamatan | Alamat
    Ditulis streaming per baris, gak numpuk di memory.

CATATAN AKURASI (v2):
    - Dedupe NPSN di tahap list (server kadang balikin baris sama 2x).
    - List API gagal -> retry lebih gigih (data vital, jgn gampang nyerah).
    - Detail API gagal (masuk PERLU_CEK_MANUAL) -> di-retry pelan di akhir
      per-provinsi, batch kecil, sebelum nyerah beneran.
"""

import asyncio
import csv
import json
import os
import sys
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HALAMAN_AWAL = ("https://data.kemendikdasmen.go.id/data-induk/satpen/daftar"
                 "?pembina=kemendikdasmen&jalur=formal&bentuk=sd,smp,sma,smk"
                 "&status=2&jenis=pendidikan-umum,pendidikan-kejuruan")

LIST_API = ("https://api.data.belajar.id/data-portal-backend/v2/master-data/satuan-pendidikan"
            "/daftar-data-induk/{kode}?pembina=kemendikdasmen&jalurPendidikan=formal"
            "&bentukPendidikan={bentuk}&statusSatuanPendidikan=2"
            "&jenisPendidikan=pendidikan-umum,pendidikan-kejuruan&limit={limit}&offset={offset}")

BENTUK_LIST = ["sd", "smp", "sma", "smk"]  # pecah per bentuk - gabungan "sd,smp,sma,smk"
                                             # kepotong lbh awal drpd meta.total (server bug)

DETAIL_API = "https://api.data.belajar.id/data-portal-backend/v1/master-data/satuan-pendidikan/details/{npsn}"
PTK_SUMMARY_API = "https://api.data.belajar.id/data-portal-backend/v1/master-data/satuan-pendidikan/details/{npsn}/ptk-summary"
PD_SUMMARY_API = "https://api.data.belajar.id/data-portal-backend/v2/master-data/satuan-pendidikan/details/{npsn}/pd-summary"

RETRY_MAKS = 4
RETRY_LIST = RETRY_MAKS * 2   # list API vital -> retry 2x lipat drpd detail
BATCH_DETAIL = 8              # 3 API/sekolah (detail+pd+ptk)
BATCH_RETRY_DETAIL = 3        # batch retry akhir, lebih kecil = lebih pelan = lebih akurat

DEBUG = "--debug" in sys.argv
TEST = "--test" in sys.argv  # cuma ambil 1 halaman list pertama per provinsi

STATS = {"sekolah": 0, "gagal_detail": 0}

KOLOM = ["NPSN", "Nama Sekolah", "Telpon", "Email", "Akreditasi", "Total Peserta Didik",
         "Total Pendidik", "Naungan", "Kabupaten", "Kecamatan", "Alamat"]

DETAIL_KOSONG = {
    "Telpon": "", "Email": "", "Akreditasi": "",
    "Total Peserta Didik": "", "Total Pendidik": "",
    "Naungan": "PERLU_CEK_MANUAL",
}

# JS snippet: fetch() dari dalem browser context, return JSON atau {__error: status}
JS_FETCH = """
async (url) => {
    try {
        const res = await fetch(url, {
            headers: {
                'accept': '*/*',
                'x-bot': '0',
                'x-bot-type': 'user',
                'x-client-app': 'data-portal-fe',
                'x-request-id': crypto.randomUUID()
            }
        });
        if (!res.ok) return {__error: res.status};
        return await res.json();
    } catch (e) {
        return {__error: String(e)};
    }
}
"""


async def fetch_json(page, url, tries=RETRY_MAKS):
    for percobaan in range(tries):
        try:
            hasil = await page.evaluate(JS_FETCH, url)
            if isinstance(hasil, dict) and "__error" in hasil:
                if DEBUG:
                    print(f"        [!] fetch gagal ({hasil['__error']}) @ {url}")
                await page.wait_for_timeout(1500 * (percobaan + 1))
                continue
            return hasil
        except Exception as e:
            if DEBUG:
                print(f"        [!] evaluate error ({e}) @ {url}")
            await page.wait_for_timeout(1500 * (percobaan + 1))
    return None


def load_provinsi():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "provinsi.json")
    if not os.path.exists(path):
        print("[!] provinsi.json gak ketemu di folder yg sama. Taruh file itu di sini dulu.")
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return dict(sorted(data.items()))


def pilih_provinsi(daftar):
    nama_list = list(daftar.keys())
    print("\n=== Pilih Provinsi ===")
    for i, nama in enumerate(nama_list, 1):
        print(f"{i:2d}. {nama}")
    print(" 0. SEMUA provinsi")
    pilihan = input("\nMasukin nomor (pisah koma kalo lebih dari 1, misal 1,3,5): ").strip()

    if pilihan == "0":
        return daftar

    idx_list = [s.strip() for s in pilihan.split(",") if s.strip()]
    hasil = {}
    for idx in idx_list:
        if idx.isdigit() and 1 <= int(idx) <= len(nama_list):
            nama = nama_list[int(idx) - 1]
            hasil[nama] = daftar[nama]
        else:
            print(f"    [!] '{idx}' diabaikan, bukan nomor valid.")
    return hasil


def bersihkan_nama_file(nama):
    import re
    nama = re.sub(r'[\\/:*?"<>|\t\n\r]', "-", nama).strip()
    return nama[:150] if nama else "tanpa_nama"


async def deteksi_limit(page, kode, bentuk):
    for coba in (100, 50, 20):
        url = LIST_API.format(kode=kode, bentuk=bentuk, limit=coba, offset=0)
        data = await fetch_json(page, url, tries=2)
        if data and data.get("meta", {}).get("limit") == coba:
            return coba
    return 20


async def ambil_satu_bentuk(page, kode, bentuk, limit, seen, offset_gagal_total):
    """Full pagination utk 1 bentuk (sd/smp/sma/smk) doang, dedupe by NPSN ke `seen` bareng.

    Query gabungan "sd,smp,sma,smk" kepotong duluan sblm nyampe meta.total (server
    bug/salah handle multi-value filter) - pecah per bentuk biar tiap query lbh
    kecil & konsisten sampe abis.
    """
    MAX_PASS = 3
    total = None

    async def satu_pass():
        nonlocal total
        offset = 0
        gagal = []
        gagal_beruntun = 0
        sebelum = len(seen)
        while True:
            if total and offset > total * 3 + 200:
                print(f"        [!] [{bentuk}] offset {offset} jauh lewatin total {total}, stop paksa (safety cap)")
                break

            url = LIST_API.format(kode=kode, bentuk=bentuk, limit=limit, offset=offset)
            data = await fetch_json(page, url, tries=RETRY_LIST)
            if not data:
                gagal.append(offset)
                gagal_beruntun += 1
                offset += limit
                if gagal_beruntun >= 5:
                    break
                continue
            gagal_beruntun = 0

            rows = data.get("data", [])
            if total is None:
                total = data.get("meta", {}).get("total", 0)
                print(f"        [{bentuk}] total: {total}")
            for r in rows:
                npsn = r.get("npsn")
                if npsn:
                    seen[npsn] = r

            offset += limit
            if TEST:
                break
            if not rows:
                break  # satu-satunya sinyal akhir data yg valid
        return gagal, len(seen) - sebelum

    pass_ke = 0
    awal = len(seen)
    gagal_terakhir = []
    for pass_ke in range(1, MAX_PASS + 1):
        gagal, baru = await satu_pass()
        gagal_terakhir = gagal  # cuma simpen dr pass paling akhir - gagal di pass
                                 # sblmnya biasanya udah ketutup pass abis ini
        print(f"        [{bentuk}] pass {pass_ke}/{MAX_PASS}: +{baru} baru, total unik {len(seen)}")
        if baru == 0 and pass_ke > 1:
            break

    tercapai = len(seen) - awal
    if total and tercapai >= total:
        gagal_terakhir = []  # datanya udah lengkap, gagal transient sblmnya gak relevan lg

    offset_gagal_total.update(
        LIST_API.format(kode=kode, bentuk=bentuk, limit=limit, offset=o) for o in gagal_terakhir
    )


async def ambil_daftar_sekolah(page, kode, limit):
    """List semua sekolah per wilayah, pecah per bentuk (sd/smp/sma/smk), dedupe by NPSN."""
    seen = {}
    offset_gagal_total = set()
    for bentuk in BENTUK_LIST:
        await ambil_satu_bentuk(page, kode, bentuk, limit, seen, offset_gagal_total)

    print(f"        total sekolah unik semua bentuk: {len(seen)}")
    if offset_gagal_total:
        print(f"        [!] {len(offset_gagal_total)} offset gagal HTTP, cek manual")

    return list(seen.values()), sorted(offset_gagal_total)


async def ambil_detail_batch(page, npsn_list):
    """Tarik detail + pd-summary + ptk-summary tiap NPSN paralel (3 API x N sekolah bareng)."""
    tugas = []
    for npsn in npsn_list:
        tugas.append(fetch_json(page, DETAIL_API.format(npsn=npsn)))
        tugas.append(fetch_json(page, PD_SUMMARY_API.format(npsn=npsn)))
        tugas.append(fetch_json(page, PTK_SUMMARY_API.format(npsn=npsn)))
    hasil_list = await asyncio.gather(*tugas)

    out = {}
    for i, npsn in enumerate(npsn_list):
        d_detail = hasil_list[i * 3]
        d_pd = hasil_list[i * 3 + 1]
        d_ptk = hasil_list[i * 3 + 2]

        if not d_detail:
            out[npsn] = None  # sinyal "gagal total", biar caller bisa retry
            continue

        sp = d_detail.get("satuanPendidikan", {})
        akreditasi = sp.get("akreditasi") or sp.get("peringkatAkreditasi") or ""

        total_pd = ""
        if d_pd:
            total_pd = (d_pd.get("rekapitulasiPesertaDidik", {})
                        .get("jumlahPesertaDidik", {}).get("total", ""))

        total_ptk = ""
        if d_ptk:
            total_ptk = (d_ptk.get("rekapitulasiPtk", {})
                         .get("jumlahPtk", {}).get("total", ""))

        out[npsn] = {
            "Telpon": sp.get("nomorTelepon", "") or "",
            "Email": sp.get("email", "") or "",
            "Akreditasi": akreditasi,
            "Total Peserta Didik": total_pd,
            "Total Pendidik": total_ptk,
            "Naungan": sp.get("namaYayasan", "") or "",
        }
        if DEBUG and i == 0:
            pd_status = 'ok' if d_pd else d_pd
            ptk_status = 'ok' if d_ptk else d_ptk
            print(f"        [debug] contoh npsn={npsn}: akreditasi_raw={akreditasi!r} pd={pd_status!r} ptk={ptk_status!r}")
    return out


def baris_dari_detail(r, detail):
    return {
        "NPSN": r.get("npsn", ""),
        "Nama Sekolah": r.get("nama", ""),
        "Telpon": detail["Telpon"],
        "Email": detail["Email"],
        "Akreditasi": detail["Akreditasi"],
        "Total Peserta Didik": detail["Total Peserta Didik"],
        "Total Pendidik": detail["Total Pendidik"],
        "Naungan": detail["Naungan"],
        "Kabupaten": r.get("namaKabupaten", ""),
        "Kecamatan": r.get("namaKecamatan", ""),
        "Alamat": r.get("alamatJalan", ""),
    }


def csv_ke_excel(path_csv, path_xlsx, kolom):
    """Baca CSV, tulis Excel rapi: header tebal warna, auto width, freeze row 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    lebar_max = [len(k) for k in kolom]

    with open(path_csv, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, 1):
            ws.append(row)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
            else:
                for c_idx, val in enumerate(row):
                    if c_idx < len(lebar_max):
                        lebar_max[c_idx] = max(lebar_max[c_idx], len(str(val)))

    for c_idx, lebar in enumerate(lebar_max, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = min(lebar + 3, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(path_xlsx)
    except PermissionError:
        import time
        print(f"    [!] '{path_xlsx}' lagi kebuka/kekunci, coba lagi 2 detik...")
        time.sleep(2)
        try:
            wb.save(path_xlsx)
        except PermissionError:
            base, ext = os.path.splitext(path_xlsx)
            path_xlsx = f"{base}_baru{ext}"
            print(f"    [!] masih kekunci, simpen ke '{path_xlsx}' sbg gantinya.")
            wb.save(path_xlsx)
    return path_xlsx


def buka_file_aman(path, mode, **kwargs):
    """Coba buka file, kalo lock (kebuka di Excel dll) retry, abis itu fallback nama baru."""
    import time
    for percobaan in range(3):
        try:
            return open(path, mode, **kwargs), path
        except PermissionError:
            print(f"    [!] '{path}' lagi kebuka/kekunci (mgkn lagi dibuka di Excel), coba lagi 2 detik...")
            time.sleep(2)
    base, ext = os.path.splitext(path)
    path_baru = f"{base}_baru{ext}"
    print(f"    [!] masih kekunci, tulis ke '{path_baru}' sbg gantinya. TUTUP file lama dulu kalo mau nama asli.")
    return open(path_baru, mode, **kwargs), path_baru


async def scrape_provinsi(page, nama_prov, kode):
    print(f"\n=== {nama_prov} (kode {kode}) ===")
    limit = await deteksi_limit(page, kode, BENTUK_LIST[0])
    print(f"    pake limit={limit} per request")

    rows, urls_gagal = await ambil_daftar_sekolah(page, kode, limit)
    print(f"    {len(rows)} sekolah unik ketemu, mulai ambil detail (telpon/email/naungan)...")

    folder = bersihkan_nama_file(nama_prov)
    os.makedirs(folder, exist_ok=True)

    if urls_gagal:
        log_path = os.path.join(folder, bersihkan_nama_file(f"{nama_prov}_list_gagal") + ".txt")
        with open(log_path, "w", encoding="utf-8") as lf:
            lf.write("\n".join(urls_gagal))
        print(f"    [!] {len(urls_gagal)} URL list gagal ditulis ke {log_path}, cek manual (buka di browser)")

    file_out = os.path.join(folder, bersihkan_nama_file(f"{nama_prov}_Swasta") + ".csv")

    f, file_out = buka_file_aman(file_out, "w", newline="", encoding="utf-8-sig")
    with f:
        writer = csv.DictWriter(f, fieldnames=KOLOM)
        writer.writeheader()

        retry_list = []
        for i in range(0, len(rows), BATCH_DETAIL):
            batch = rows[i:i + BATCH_DETAIL]
            npsn_list = [r["npsn"] for r in batch]
            detail_map = await ambil_detail_batch(page, npsn_list)

            for r in batch:
                detail = detail_map.get(r["npsn"])
                if detail is None:
                    retry_list.append(r)  # gagal total -> coba lagi di akhir
                    continue
                writer.writerow(baris_dari_detail(r, detail))
            f.flush()
            STATS["sekolah"] += len(batch)
            print(f"        {min(i + BATCH_DETAIL, len(rows))}/{len(rows)} sekolah tercatat...")

        if retry_list:
            print(f"    retry {len(retry_list)} sekolah gagal detail, batch kecil pelan-pelan...")
            masih_gagal = []
            for i in range(0, len(retry_list), BATCH_RETRY_DETAIL):
                batch = retry_list[i:i + BATCH_RETRY_DETAIL]
                npsn_list = [r["npsn"] for r in batch]
                detail_map = await ambil_detail_batch(page, npsn_list)
                for r in batch:
                    detail = detail_map.get(r["npsn"])
                    if detail is None:
                        masih_gagal.append(r)
                        detail = DETAIL_KOSONG
                        STATS["gagal_detail"] += 1
                    writer.writerow(baris_dari_detail(r, detail))
                f.flush()
            if masih_gagal:
                print(f"    [!] {len(masih_gagal)} sekolah tetep gagal detail stlh retry -> Naungan='PERLU_CEK_MANUAL'")

    file_xlsx = os.path.join(folder, bersihkan_nama_file(f"{nama_prov}_Swasta") + ".xlsx")
    file_xlsx = csv_ke_excel(file_out, file_xlsx, KOLOM)

    print(f"=== {nama_prov} SELESAI -> {file_out} + {file_xlsx} ({len(rows)} baris) ===")


async def main():
    daftar = load_provinsi()
    daftar = {k: v for k, v in daftar.items() if v}
    terpilih = pilih_provinsi(daftar)

    if not terpilih:
        print("Gak ada provinsi valid dipilih, keluar.")
        return

    print(f"\nMulai scrape {len(terpilih)} provinsi: {', '.join(terpilih.keys())}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=not DEBUG)
        page = await browser.new_page()
        # buka halaman aslinya dulu, biar origin/cookie/context kebentuk bener
        await page.goto(HALAMAN_AWAL, wait_until="networkidle", timeout=90000)
        await page.wait_for_timeout(1000)

        for nama_prov, kode in terpilih.items():
            await scrape_provinsi(page, nama_prov, kode)

        await browser.close()

    print(f"\nSemua selesai. Total {STATS['sekolah']} sekolah, "
          f"{STATS['gagal_detail']} gagal ambil detail (cek kolom Naungan='PERLU_CEK_MANUAL').")


if __name__ == "__main__":
    asyncio.run(main())